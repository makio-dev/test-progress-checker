#!/usr/bin/env python3
"""
ITB- Excel ファイル進捗チェッカー
指定フォルダ配下の ITB- で始まる Excel ファイル内の ITB- で始まるシートを検証し、
結果を Excel ファイルに出力する。
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime

import openpyxl


def get_cell_value(ws, row, col):
    """セルの値を取得する（関数の場合は None 扱いにしない）"""
    cell = ws.cell(row=row, column=col)
    return cell.value


def get_cell_raw_value(ws, row, col):
    """セルの値を取得する。関数（=で始まる）の場合は None を返す"""
    cell = ws.cell(row=row, column=col)
    val = cell.value
    if isinstance(val, str) and val.strip().startswith("="):
        return None
    return val


def is_empty(val):
    """値が空かどうかを判定"""
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    return False


def col_letter_to_num(letter):
    """列文字を列番号に変換 (A=1, B=2, ...)"""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def validate_sheet(ws, sheet_name, file_path, file_name):
    """シートを検証してエラーリストを返す"""
    errors = []

    def add_error(msg):
        errors.append({
            "error": msg,
            "sheet": sheet_name,
            "file_name": file_name,
            "full_path": file_path,
        })

    # D5: シナリオ番号
    if is_empty(get_cell_value(ws, 5, col_letter_to_num("D"))):
        add_error("シナリオ番号が未入力です")

    # E5: タイトル
    if is_empty(get_cell_value(ws, 5, col_letter_to_num("E"))):
        add_error("タイトルが未入力です")

    # E8: 概要
    if is_empty(get_cell_value(ws, 8, col_letter_to_num("E"))):
        add_error("概要が未入力です")

    # T5: テストケース作成日
    if is_empty(get_cell_value(ws, 5, col_letter_to_num("T"))):
        add_error("テストケース作成日が未入力です")

    # T6: テストケース作成者
    if is_empty(get_cell_value(ws, 6, col_letter_to_num("T"))):
        add_error("テストケース作成者が未入力です")

    # T7: テスト実施者
    if is_empty(get_cell_value(ws, 7, col_letter_to_num("T"))):
        add_error("テスト実施者が未入力です")

    # T8: テスト検証者
    if is_empty(get_cell_value(ws, 8, col_letter_to_num("T"))):
        add_error("テスト検証者が未入力です")

    # 19行目以降のデータ検証
    col_c = col_letter_to_num("C")
    col_d = col_letter_to_num("D")
    col_v = col_letter_to_num("V")
    col_o = col_letter_to_num("O")
    col_q = col_letter_to_num("Q")
    col_s = col_letter_to_num("S")

    max_row = ws.max_row or 18
    for row in range(19, max_row + 1):
        c_val = get_cell_value(ws, row, col_c)
        c_has_value = not is_empty(c_val)

        # D~V列に値があるか確認
        dv_has_value = False
        for col in range(col_d, col_v + 1):
            if not is_empty(get_cell_value(ws, row, col)):
                dv_has_value = True
                break

        # C列に値がない行はスキップ
        if not c_has_value:
            continue

        # C列に値があって D~V に値がない → 検証内容のみ記載のケースあり
        if not dv_has_value:
            add_error(f"{row}行目: 検証内容のみ記載のケースがあります")

        # O列の値を取得（ハイフンかどうかで後続チェックが変わる）
        o_val = get_cell_value(ws, row, col_o)
        is_excluded = isinstance(o_val, str) and o_val.strip() == "-"

        # C列に値がある場合、Q列(実施日予定)・S列(検証日予定)が必須（O列がハイフンの場合は除外）
        if c_has_value and not is_excluded:
            q_val = get_cell_value(ws, row, col_q)
            s_val = get_cell_value(ws, row, col_s)
            if is_empty(q_val):
                add_error(f"{row}行目: ケースIDがあるのに実施日（予定）が未入力です")
            if is_empty(s_val):
                add_error(f"{row}行目: ケースIDがあるのに検証日（予定）が未入力です")

        # C列に値がありO列がハイフン(-)の場合、V列も必須
        if c_has_value and is_excluded:
            if isinstance(o_val, str) and o_val.strip() == "-":
                v_val = get_cell_value(ws, row, col_v)
                if is_empty(v_val):
                    add_error(f"{row}行目: 実施対象外ケースの場合は欠陥内容／備考欄に理由を記載してください")

    return errors


def validate_file(file_path):
    """Excelファイルを検証してエラーリストを返す"""
    errors = []
    file_name = os.path.basename(file_path)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return [{
            "error": f"ファイルを開けません: {e}",
            "sheet": "-",
            "file_name": file_name,
            "full_path": file_path,
        }]

    # ITB- で始まるシートを収集
    itb_sheets = [name for name in wb.sheetnames if name.startswith("ITB-")]

    # シートが「ITB-」から始まるシートが１件しか存在しないこと
    if len(itb_sheets) == 0:
        errors.append({
            "error": "ITB- で始まるシートが存在しません",
            "sheet": "-",
            "file_name": file_name,
            "full_path": file_path,
        })
    elif len(itb_sheets) > 1:
        errors.append({
            "error": f"ITB- で始まるシートが {len(itb_sheets)} 件存在します（1件のみ許可）: {', '.join(itb_sheets)}",
            "sheet": "-",
            "file_name": file_name,
            "full_path": file_path,
        })

    # 各 ITB- シートを検証
    for sheet_name in itb_sheets:
        ws = wb[sheet_name]
        sheet_errors = validate_sheet(ws, sheet_name, file_path, file_name)
        errors.extend(sheet_errors)

    wb.close()
    return errors


def find_itb_files(folder_path):
    """フォルダ配下の ITB- で始まる Excel ファイルを再帰的に探す"""
    files = []
    for root, dirs, filenames in os.walk(folder_path):
        # 隠しフォルダや一時ファイルをスキップ
        dirs[:] = [d for d in dirs if not d.startswith(".") and not d.startswith("~")]
        for f in filenames:
            if f.startswith("ITB-") and (f.endswith(".xlsx") or f.endswith(".xlsm")) and not f.startswith("~$"):
                files.append(os.path.join(root, f))
    return files


def write_results(errors, output_path):
    """結果を Excel ファイルに出力"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "検証結果"

    # ヘッダー
    headers = ["No.", "エラー内容", "シート名", "ファイル名", "フルパス"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # データ
    for idx, err in enumerate(errors, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=err["error"])
        ws.cell(row=idx + 1, column=3, value=err["sheet"])
        ws.cell(row=idx + 1, column=4, value=err["file_name"])
        ws.cell(row=idx + 1, column=5, value=err["full_path"])

    # 列幅調整
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 80

    wb.save(output_path)
    wb.close()


def main():
    root = tk.Tk()
    root.withdraw()

    # フォルダ選択ダイアログ
    folder_path = filedialog.askdirectory(title="検証対象フォルダを選択してください")
    if not folder_path:
        messagebox.showinfo("キャンセル", "フォルダが選択されませんでした。")
        sys.exit(0)

    # ITB- ファイル検索
    files = find_itb_files(folder_path)
    if not files:
        messagebox.showinfo("結果", f"対象フォルダ内に ITB- で始まる Excel ファイルが見つかりませんでした。\n{folder_path}")
        sys.exit(0)

    # 全ファイル検証
    all_errors = []
    for file_path in files:
        file_errors = validate_file(file_path)
        all_errors.extend(file_errors)

    # 結果出力
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(folder_path, f"検証結果_{timestamp}.xlsx")
    write_results(all_errors, output_path)

    if all_errors:
        messagebox.showinfo(
            "完了",
            f"検証が完了しました。\n\n"
            f"対象ファイル数: {len(files)}\n"
            f"エラー件数: {len(all_errors)}\n\n"
            f"結果ファイル:\n{output_path}"
        )
    else:
        messagebox.showinfo(
            "完了",
            f"検証が完了しました。エラーはありませんでした。\n\n"
            f"対象ファイル数: {len(files)}\n\n"
            f"結果ファイル:\n{output_path}"
        )


if __name__ == "__main__":
    main()
